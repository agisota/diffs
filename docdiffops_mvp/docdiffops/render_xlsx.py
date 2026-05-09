from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
RED_FILL = PatternFill("solid", fgColor="F8CBAD")
GREEN_FILL = PatternFill("solid", fgColor="C6E0B4")
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
GRAY_FILL = PatternFill("solid", fgColor="E7E6E6")
BORDER = Border(bottom=Side(style="thin", color="D9E2F3"))


def style_header(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = BORDER
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def autofit(ws, max_width=70):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            val = "" if cell.value is None else str(cell.value)
            widths[cell.column] = max(widths.get(cell.column, 0), min(max_width, len(val) + 2))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(max_width, width))


def event_row(event: dict[str, Any]) -> list[Any]:
    lhs = event.get("lhs") or {}
    rhs = event.get("rhs") or {}
    return [
        event.get("event_id"),
        event.get("pair_id"),
        event.get("comparison_type"),
        event.get("status"),
        event.get("severity"),
        event.get("confidence"),
        event.get("score"),
        event.get("review_required"),
        event.get("lhs_doc_id"),
        lhs.get("doc_title"),
        lhs.get("page_no"),
        lhs.get("block_id"),
        str(lhs.get("bbox")) if lhs.get("bbox") else "",
        lhs.get("quote"),
        event.get("rhs_doc_id"),
        rhs.get("doc_title"),
        rhs.get("page_no"),
        rhs.get("block_id"),
        str(rhs.get("bbox")) if rhs.get("bbox") else "",
        rhs.get("quote"),
        event.get("explanation_short"),
        event.get("reviewer_decision"),
        event.get("reviewer_comment"),
    ]


def fill_for_status(status: str):
    if status == "deleted":
        return RED_FILL
    if status == "added":
        return GREEN_FILL
    if status in {"partial", "modified"}:
        return YELLOW_FILL
    return GRAY_FILL


def render_evidence_matrix(out_path: Path, state: dict[str, Any], all_events: list[dict[str, Any]], pair_summaries: list[dict[str, Any]]) -> Path:
    wb = Workbook()

    ws = wb.active
    ws.title = "00_summary"
    ws.append(["metric", "value"])
    metrics = {
        "batch_id": state["batch_id"],
        "title": state.get("title"),
        "documents": len(state.get("documents", [])),
        "pairs": len(pair_summaries),
        "events_total": len(all_events),
        "high": sum(1 for e in all_events if e.get("severity") == "high"),
        "review_required": sum(1 for e in all_events if e.get("review_required")),
        "added": sum(1 for e in all_events if e.get("status") == "added"),
        "deleted": sum(1 for e in all_events if e.get("status") == "deleted"),
        "partial": sum(1 for e in all_events if e.get("status") == "partial"),
        "modified": sum(1 for e in all_events if e.get("status") == "modified"),
        # PR-2.1: surface cache effectiveness in the evidence matrix.
        "cache_extract_hits": sum(
            1 for d in state.get("documents", []) if d.get("cache_extract_hit")
        ),
        "cache_compare_hits": sum(
            1 for p in (pair_summaries or []) if p.get("cache_hit")
        ),
    }
    for k, v in metrics.items():
        ws.append([k, v])
    style_header(ws)
    autofit(ws)

    ws = wb.create_sheet("01_source_inventory")
    ws.append(["doc_id", "title", "filename", "doc_type", "source_rank", "sha256", "raw_path", "canonical_pdf", "extracted_json"])
    for d in state.get("documents", []):
        ws.append([d.get("doc_id"), d.get("title"), d.get("filename"), d.get("doc_type"), d.get("source_rank"), d.get("sha256"), d.get("raw_path"), d.get("canonical_pdf"), d.get("extracted_json")])
    style_header(ws)
    autofit(ws)

    ws = wb.create_sheet("02_pair_matrix")
    ws.append(["pair_id", "lhs_doc_id", "rhs_doc_id", "events_total", "same_count", "partial_count", "added_count", "deleted_count", "high_count", "review_required_count"])
    for s in pair_summaries:
        ws.append([s.get(k) for k in ["pair_id", "lhs_doc_id", "rhs_doc_id", "events_total", "same_count", "partial_count", "added_count", "deleted_count", "high_count", "review_required_count"]])
    style_header(ws)
    autofit(ws)

    headers = [
        "event_id", "pair_id", "comparison_type", "status", "severity", "confidence", "score", "review_required",
        "lhs_doc_id", "lhs_title", "lhs_page", "lhs_block_id", "lhs_bbox", "lhs_quote",
        "rhs_doc_id", "rhs_title", "rhs_page", "rhs_block_id", "rhs_bbox", "rhs_quote",
        "explanation_short", "reviewer_decision", "reviewer_comment",
    ]
    ws = wb.create_sheet("03_diff_events_all")
    ws.append(headers)
    for e in all_events:
        ws.append(event_row(e))
        fill = fill_for_status(e.get("status"))
        for cell in ws[ws.max_row]:
            cell.fill = fill
    style_header(ws)
    autofit(ws)

    for sheet_name, predicate in [
        ("04_high_risk", lambda e: e.get("severity") == "high"),
        ("05_partial_matches", lambda e: e.get("status") == "partial"),
        ("06_review_queue", lambda e: e.get("review_required")),
        ("07_added_deleted", lambda e: e.get("status") in {"added", "deleted"}),
    ]:
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        for e in filter(predicate, all_events):
            ws.append(event_row(e))
            fill = fill_for_status(e.get("status"))
            for cell in ws[ws.max_row]:
                cell.fill = fill
        style_header(ws)
        autofit(ws)

    # PR-2.1: status × severity pivot.
    ws = wb.create_sheet("08_status_severity")
    statuses = ["same", "partial", "modified", "added", "deleted", "contradicts", "manual_review"]
    severities = ["high", "medium", "low"]
    ws.append(["status \\ severity", *severities, "total"])
    for st in statuses:
        row = [st]
        total = 0
        for sv in severities:
            n = sum(
                1 for e in all_events
                if (e.get("status") or "") == st and (e.get("severity") or "low") == sv
            )
            row.append(n)
            total += n
        row.append(total)
        ws.append(row)
    style_header(ws)
    autofit(ws)

    # PR-2.1: by-source-rank sheet — events grouped by lhs/rhs rank pair.
    docs_by_id = {d.get("doc_id"): d for d in state.get("documents", [])}
    rank_pairs: dict[tuple[int, int], int] = {}
    for ev in all_events:
        lr = int((docs_by_id.get(ev.get("lhs_doc_id"), {}) or {}).get("source_rank") or 3)
        rr = int((docs_by_id.get(ev.get("rhs_doc_id"), {}) or {}).get("source_rank") or 3)
        key = (min(lr, rr), max(lr, rr))
        rank_pairs[key] = rank_pairs.get(key, 0) + 1
    ws = wb.create_sheet("09_by_source_rank")
    ws.append(["rank_lo", "rank_hi", "label", "events"])
    label_map = {1: "official_NPA", 2: "departmental", 3: "analytics"}
    for (lo, hi), n in sorted(rank_pairs.items()):
        ws.append([lo, hi, f"{label_map[lo]} ↔ {label_map[hi]}", n])
    style_header(ws)
    autofit(ws)

    # PR-2.1: cache_metrics sheet.
    ws = wb.create_sheet("10_cache_metrics")
    docs = state.get("documents", [])
    pairs = pair_summaries or []
    extract_hits = sum(1 for d in docs if d.get("cache_extract_hit"))
    compare_hits = sum(1 for p in pairs if p.get("cache_hit"))
    ws.append(["scope", "hits", "total", "ratio"])
    ws.append([
        "extract", extract_hits, len(docs),
        round(extract_hits / len(docs), 3) if docs else 0,
    ])
    ws.append([
        "compare", compare_hits, len(pairs),
        round(compare_hits / len(pairs), 3) if pairs else 0,
    ])
    style_header(ws)
    autofit(ws)

    # 11_topic_clusters: cross-pair clustering. Falls back to in-line
    # cluster_events when state doesn't carry the precomputed list.
    try:
        clusters = state.get("topic_clusters")
        if clusters is None:
            from .legal.cross_pair import cluster_events as _ce
            clusters = _ce(all_events)
        ws = wb.create_sheet("11_topic_clusters")
        ws.append([
            "cluster_id", "topic", "status", "severity", "count",
            "pair_count", "comparison_types", "explanations",
        ])
        for c in clusters or []:
            ws.append([
                c.get("cluster_id", ""),
                c.get("topic", ""),
                c.get("status", ""),
                c.get("severity", ""),
                int(c.get("count", 0)),
                len(c.get("pair_ids", []) or []),
                ", ".join(c.get("comparison_types", []) or []),
                " | ".join(c.get("explanations", []) or []),
            ])
            fill = fill_for_status(c.get("status"))
            for cell in ws[ws.max_row]:
                cell.fill = fill
        style_header(ws)
        autofit(ws)
    except Exception:
        pass

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
