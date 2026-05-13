"""Final consolidated Russian XLSX summary for DocDiffOps v10.

17 sheets, all numbered and named in Russian, with proper text wrapping,
color-coded status cells, frozen header rows, and color-blind dual-coding.

Output: migration_v10_out/Сводный_отчёт_v10.xlsx
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter

# Import data
sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
from docdiffops.forensic_render import PALETTE, STATUS_RU, STATUS_PALETTE  # noqa: E402
from docdiffops.forensic import STATUS_TO_MARK  # noqa: E402
from scripts.presentation_v10.data_loader import load_data  # noqa: E402

REPO_ROOT = Path(__file__).resolve().parents[3]
OUT_PATH = REPO_ROOT / "migration_v10_out" / "Сводный_отчёт_v10.xlsx"

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
BODY_FONT = Font(name="Arial", size=11)
BODY_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
ZEBRA_FILL = PatternFill("solid", fgColor="F9FAFB")

STATUS_FILLS: dict[str, PatternFill] = {
    "match":           PatternFill("solid", fgColor="DCFCE7"),
    "partial_overlap": PatternFill("solid", fgColor="FEF3C7"),
    "contradiction":   PatternFill("solid", fgColor="FEE2E2"),
    "outdated":        PatternFill("solid", fgColor="DBEAFE"),
    "source_gap":      PatternFill("solid", fgColor="EDE9FE"),
    "manual_review":   PatternFill("solid", fgColor="FFEDD5"),
    "not_comparable":  PatternFill("solid", fgColor="F3F4F6"),
}

PRIORITY_FILLS: dict[str, PatternFill] = {
    "P0": PatternFill("solid", fgColor="FEE2E2"),
    "P1": PatternFill("solid", fgColor="FEF3C7"),
    "P2": PatternFill("solid", fgColor="CCFBF1"),
}

SEVERITY_FILLS: dict[str, PatternFill] = {
    "high":   PatternFill("solid", fgColor="FEE2E2"),
    "medium": PatternFill("solid", fgColor="FFEDD5"),
    "low":    PatternFill("solid", fgColor="DBEAFE"),
}

QA_FILLS: dict[str, PatternFill] = {
    "PASS": PatternFill("solid", fgColor="DCFCE7"),
    "WARN": PatternFill("solid", fgColor="FEF3C7"),
    "FAIL": PatternFill("solid", fgColor="FEE2E2"),
}

RELATION_FILLS: dict[str, PatternFill] = {
    "amends":     PatternFill("solid", fgColor="DBEAFE"),
    "supersedes": PatternFill("solid", fgColor="FEE2E2"),
    "references": PatternFill("solid", fgColor="F3F4F6"),
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _style_header_row(ws, row: int, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = HEADER_ALIGN
        c.border = BORDER
    ws.row_dimensions[row].height = 32


def _style_body_cell(
    cell,
    *,
    status_eng: str | None = None,
    zebra: bool = False,
    fill: PatternFill | None = None,
) -> None:
    cell.font = BODY_FONT
    cell.alignment = BODY_ALIGN
    cell.border = BORDER
    if fill is not None:
        cell.fill = fill
    elif status_eng and status_eng in STATUS_FILLS:
        cell.fill = STATUS_FILLS[status_eng]
    elif zebra:
        cell.fill = ZEBRA_FILL


def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(w, 60)


def _add_status_text(status_eng: str) -> str:
    glyph = STATUS_TO_MARK.get(status_eng, "")
    label = STATUS_RU.get(status_eng, status_eng)
    return f"{glyph} {label}"


def _write_row(
    ws,
    row: int,
    values: list,
    *,
    zebra: bool = False,
    status_col: int | None = None,
    status_eng: str | None = None,
    fill_map: dict[int, PatternFill] | None = None,
) -> None:
    """Write a row of values with uniform body styling."""
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        fill = None
        if fill_map and col in fill_map:
            fill = fill_map[col]
        elif status_col is not None and col == status_col and status_eng:
            fill = STATUS_FILLS.get(status_eng)
        _style_body_cell(c, zebra=zebra, fill=fill)


def _freeze(ws, cell: str = "A2") -> None:
    ws.freeze_panes = cell


# ---------------------------------------------------------------------------
# Sheet 01: Сводка
# ---------------------------------------------------------------------------

def build_sheet_01_summary(wb: Workbook, data) -> None:
    ws = wb.create_sheet("01 Сводка")

    # Title banner
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "DocDiffOps v10 — Сводный отчёт"
    c.font = Font(name="Arial", size=18, bold=True, color="065A82")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor="EFF6FF")
    c.border = BORDER
    ws.row_dimensions[1].height = 40

    # KPI labels row 3
    cn = data.control_numbers
    qa = data.qa
    kpi_items = [
        ("Документов", str(cn.get("documents", 27)), "065A82"),
        ("Пар сравнения", str(cn.get("pairs", 351)), "1C7293"),
        ("Diff-событий", str(cn.get("events", 312)), "21295C"),
        ("QA-гейт", f"{qa.get('passed',12)}/{qa.get('total',12)} PASS", "16A34A"),
        ("Версия", "10.0.0", "374151"),
        ("Дата", "2026-05-09", "374151"),
    ]
    for i, (label, value, color) in enumerate(kpi_items, 1):
        lc = ws.cell(row=3, column=i, value=label)
        lc.font = Font(name="Arial", size=9, color="6B7280")
        lc.alignment = Alignment(horizontal="center")
        lc.border = BORDER

        vc = ws.cell(row=4, column=i, value=value)
        vc.font = Font(name="Arial", size=16, bold=True, color=color)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.fill = PatternFill("solid", fgColor="F8FAFC")
        vc.border = BORDER
    ws.row_dimensions[4].height = 30

    # Section: Pairs by status
    ws.cell(row=6, column=1, value="Распределение пар по статусам").font = Font(
        name="Arial", size=13, bold=True, color="1F2937"
    )
    headers_p = ["Статус", "Количество", "Доля, %"]
    for col, h in enumerate(headers_p, 1):
        ws.cell(row=7, column=col, value=h)
    _style_header_row(ws, 7, len(headers_p))

    pairs_status = data.pairs_by_status()
    total_pairs = sum(pairs_status.values()) or 1
    row = 8
    for status_eng in [
        "match", "partial_overlap", "contradiction", "outdated",
        "source_gap", "manual_review", "not_comparable"
    ]:
        count = pairs_status.get(status_eng, 0)
        pct = round(count / total_pairs * 100, 1)
        zebra = (row % 2 == 0)
        for col, val in enumerate([_add_status_text(status_eng), count, pct], 1):
            c = ws.cell(row=row, column=col, value=val)
            if col == 1:
                _style_body_cell(c, status_eng=status_eng)
            else:
                _style_body_cell(c, zebra=zebra)
        row += 1
    # Total row
    tc1 = ws.cell(row=row, column=1, value="Итого")
    tc1.font = Font(name="Arial", size=11, bold=True)
    tc1.border = BORDER
    tc2 = ws.cell(row=row, column=2, value=f"=SUM(B8:B{row-1})")
    tc2.font = Font(name="Arial", size=11, bold=True)
    tc2.border = BORDER
    row += 2

    # Section: Events by status
    ws.cell(row=row, column=1, value="Распределение событий по статусам").font = Font(
        name="Arial", size=13, bold=True, color="1F2937"
    )
    row += 1
    event_hdr_row = row
    for col, h in enumerate(["Статус", "Количество", "Доля, %"], 1):
        ws.cell(row=row, column=col, value=h)
    _style_header_row(ws, row, 3)
    row += 1

    ev_status = data.status_distribution()
    total_ev = sum(ev_status.values()) or 1
    ev_data_start = row
    for status_eng in [
        "match", "partial_overlap", "contradiction", "outdated",
        "source_gap", "manual_review", "not_comparable"
    ]:
        count = ev_status.get(status_eng, 0)
        pct = round(count / total_ev * 100, 1)
        zebra = (row % 2 == 0)
        for col, val in enumerate([_add_status_text(status_eng), count, pct], 1):
            c = ws.cell(row=row, column=col, value=val)
            if col == 1:
                _style_body_cell(c, status_eng=status_eng)
            else:
                _style_body_cell(c, zebra=zebra)
        row += 1
    tc1 = ws.cell(row=row, column=1, value="Итого")
    tc1.font = Font(name="Arial", size=11, bold=True)
    tc1.border = BORDER
    tc2 = ws.cell(row=row, column=2, value=f"=SUM(B{ev_data_start}:B{row-1})")
    tc2.font = Font(name="Arial", size=11, bold=True)
    tc2.border = BORDER
    row += 2

    # Footer note
    note = ws.cell(
        row=row, column=1,
        value="Содержание других листов см. лист 02 Легенда"
    )
    note.font = Font(name="Arial", size=10, italic=True, color="6B7280")

    _set_col_widths(ws, [32, 18, 12, 18, 18, 18])
    _freeze(ws, "A2")


# ---------------------------------------------------------------------------
# Sheet 02: Легенда
# ---------------------------------------------------------------------------

def build_sheet_02_legend(wb: Workbook, data) -> None:  # noqa: ARG001
    ws = wb.create_sheet("02 Легенда")

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "Легенда статусов и символов"
    c.font = Font(name="Arial", size=14, bold=True, color="1F2937")
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = BORDER
    ws.row_dimensions[1].height = 28

    # Status table header
    headers = ["Глиф", "Статус (ru)", "Статус (en)", "Цвет (hex)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    _style_header_row(ws, 2, 4)

    statuses_order = [
        "match", "partial_overlap", "contradiction", "outdated",
        "source_gap", "manual_review", "not_comparable"
    ]
    hex_map = {
        "match": "#DCFCE7 / #16A34A",
        "partial_overlap": "#FEF3C7 / #F59E0B",
        "contradiction": "#FEE2E2 / #DC2626",
        "outdated": "#DBEAFE / #2563EB",
        "source_gap": "#EDE9FE / #7C3AED",
        "manual_review": "#FFEDD5 / #EA580C",
        "not_comparable": "#F3F4F6 / #6B7280",
    }
    for i, status_eng in enumerate(statuses_order):
        row = i + 3
        glyph = STATUS_TO_MARK.get(status_eng, "")
        ru = STATUS_RU.get(status_eng, status_eng)
        hex_val = hex_map.get(status_eng, "")
        zebra = (row % 2 == 0)
        for col, val in enumerate([glyph, ru, status_eng, hex_val], 1):
            c = ws.cell(row=row, column=col, value=val)
            _style_body_cell(c, status_eng=status_eng, zebra=zebra)

    # Source ranks
    row = 12
    ws.cell(row=row, column=1, value="Ранги источников").font = Font(
        name="Arial", size=13, bold=True, color="1F2937"
    )
    row += 1
    for col, h in enumerate(["Ранг", "Описание", "Пример"], 1):
        ws.cell(row=row, column=col, value=h)
    _style_header_row(ws, row, 3)
    row += 1
    rank_data = [
        ("1", "Официальные НПА (федеральные законы, указы, постановления)", "ФЗ-115, Указ №1 ,ПП №2573"),
        ("2", "Ведомственные документы (приказы, брошюры министерств)", "Брошюра Минэка, приказ МВД"),
        ("3", "Аналитика, презентации, дайджесты", "ВЦИОМ, Клерк, Нейрон"),
    ]
    for rank, desc, ex in rank_data:
        zebra = (row % 2 == 0)
        for col, val in enumerate([rank, desc, ex], 1):
            c = ws.cell(row=row, column=col, value=val)
            _style_body_cell(c, zebra=zebra)
        row += 1

    # Glyph explanation
    row += 1
    ws.cell(row=row, column=1, value="Расшифровка глифов").font = Font(
        name="Arial", size=13, bold=True, color="1F2937"
    )
    row += 1
    note_text = (
        "Глифы ✓ ≈ ⚠ ↻ ∅ ? — используются как dual-coding для "
        "людей с нарушениями цветовосприятия. Цвет фона несёт ту же "
        "информацию, что и глиф: смысл считывается любым из каналов."
    )
    c = ws.cell(row=row, column=1, value=note_text)
    c.font = Font(name="Arial", size=10, italic=True, color="374151")
    c.alignment = Alignment(wrap_text=True)
    ws.merge_cells(f"A{row}:D{row}")
    ws.row_dimensions[row].height = 50

    _set_col_widths(ws, [8, 30, 25, 30])
    _freeze(ws, "A3")


# ---------------------------------------------------------------------------
# Sheet 03: Документы
# ---------------------------------------------------------------------------

def build_sheet_03_documents(wb: Workbook, data) -> None:
    ws = wb.create_sheet("03 Документы")

    headers = ["ID", "Код", "Ранг", "Заголовок", "Тип", "URL / Источник"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    docs = sorted(data.documents, key=lambda d: d.get("id", ""))
    for i, doc in enumerate(docs):
        row = i + 2
        zebra = (i % 2 == 0)
        vals = [
            doc.get("id", ""),
            doc.get("code", ""),
            doc.get("rank", ""),
            doc.get("title", ""),
            doc.get("type", ""),
            doc.get("url", ""),
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            _style_body_cell(c, zebra=zebra)

    # Summary formulas below data
    formula_row = len(docs) + 2 + 1
    ws.cell(row=formula_row, column=1, value="Ранг 1 (НПА):")
    ws.cell(row=formula_row, column=1).font = Font(name="Arial", size=10, bold=True, color="6B7280")
    ws.cell(row=formula_row, column=3, value=f'=COUNTIF(C2:C{len(docs)+1},"1")')
    ws.cell(row=formula_row, column=3).font = Font(name="Arial", size=10, color="16A34A")

    ws.cell(row=formula_row + 1, column=1, value="Ранг 2 (ведомственные):")
    ws.cell(row=formula_row + 1, column=1).font = Font(name="Arial", size=10, bold=True, color="6B7280")
    ws.cell(row=formula_row + 1, column=3, value=f'=COUNTIF(C2:C{len(docs)+1},"2")')
    ws.cell(row=formula_row + 1, column=3).font = Font(name="Arial", size=10, color="F59E0B")

    ws.cell(row=formula_row + 2, column=1, value="Ранг 3 (аналитика):")
    ws.cell(row=formula_row + 2, column=1).font = Font(name="Arial", size=10, bold=True, color="6B7280")
    ws.cell(row=formula_row + 2, column=3, value=f'=COUNTIF(C2:C{len(docs)+1},"3")')
    ws.cell(row=formula_row + 2, column=3).font = Font(name="Arial", size=10, color="6B7280")

    _set_col_widths(ws, [8, 22, 6, 45, 30, 50])
    _freeze(ws, "A2")


# ---------------------------------------------------------------------------
# Sheet 04: Темы — каталог
# ---------------------------------------------------------------------------

def build_sheet_04_themes_catalog(wb: Workbook, data) -> None:
    ws = wb.create_sheet("04 Темы — каталог")

    headers = [
        "ID темы", "Название темы", "Документов в теме",
        "Событий", "Задач в очереди", "Описание / комментарий"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    # Build unique themes from correlation_matrix which has one row per theme
    themes_seen: dict[str, str] = {}
    for row_data in data.correlation_matrix:
        tid = row_data.get("theme_id", "")
        tname = row_data.get("theme_name", "")
        if tid:
            themes_seen[tid] = tname

    # Count docs per theme from theme_doc
    docs_per_theme: dict[str, int] = {}
    for r in data.theme_doc:
        tid = r.get("theme_id", "")
        role = r.get("role", "")
        if tid and role != "не покрывает":
            docs_per_theme[tid] = docs_per_theme.get(tid, 0) + 1

    # Count events per theme
    events_per_theme: dict[str, int] = {}
    for ev in data.events_all:
        tid = ev.get("theme_id", "")
        if tid:
            events_per_theme[tid] = events_per_theme.get(tid, 0) + 1

    # Count review queue tasks per theme
    queue_per_theme: dict[str, int] = {}
    for rq in data.review_queue:
        theme_name = rq.get("theme", "")
        queue_per_theme[theme_name] = queue_per_theme.get(theme_name, 0) + 1

    # Sort themes: T01..T17 then T00
    def _theme_sort_key(tid: str) -> int:
        try:
            return int(tid[1:]) if tid.startswith("T") else 999
        except ValueError:
            return 999

    sorted_themes = sorted(themes_seen.keys(), key=_theme_sort_key)

    for i, tid in enumerate(sorted_themes):
        row = i + 2
        zebra = (i % 2 == 0)
        tname = themes_seen[tid]
        doc_count = docs_per_theme.get(tid, 0)
        ev_count = events_per_theme.get(tid, 0)
        # match queue by theme name
        rq_count = queue_per_theme.get(tname, 0)
        vals = [tid, tname, doc_count, ev_count, rq_count, ""]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            _style_body_cell(c, zebra=zebra)

    _set_col_widths(ws, [10, 40, 18, 12, 16, 50])
    _freeze(ws, "A2")


# ---------------------------------------------------------------------------
# Sheet 05: Корреляция темы × документы
# ---------------------------------------------------------------------------

def build_sheet_05_correlation(wb: Workbook, data) -> None:
    ws = wb.create_sheet("05 Корреляция тема×документ")

    doc_ids = [f"D{i:02d}" for i in range(1, 28)]  # D01..D27

    # Header row
    ws.cell(row=1, column=1, value="ID темы")
    ws.cell(row=1, column=2, value="Название темы")
    for j, did in enumerate(doc_ids, 3):
        ws.cell(row=1, column=j, value=did)
    _style_header_row(ws, 1, 2 + len(doc_ids))

    # Build a lookup from correlation_matrix
    corr_lookup: dict[str, dict[str, int]] = {}
    for row_data in data.correlation_matrix:
        tid = row_data.get("theme_id", "")
        tname = row_data.get("theme_name", "")
        corr_lookup[tid] = {"theme_name": tname}  # type: ignore[assignment]
        for did in doc_ids:
            try:
                corr_lookup[tid][did] = int(row_data.get(did, "0"))
            except ValueError:
                corr_lookup[tid][did] = 0

    def _theme_sort_key(tid: str) -> int:
        try:
            return int(tid[1:]) if tid.startswith("T") else 999
        except ValueError:
            return 999

    sorted_themes = sorted(corr_lookup.keys(), key=_theme_sort_key)
    PRESENT_FILL = PatternFill("solid", fgColor="1C7293")
    ABSENT_FILL = PatternFill("solid", fgColor="F8FAFC")

    for i, tid in enumerate(sorted_themes):
        row = i + 2
        zebra = (i % 2 == 0)
        row_data = corr_lookup[tid]
        tname = row_data.get("theme_name", "")  # type: ignore[call-overload]

        c1 = ws.cell(row=row, column=1, value=tid)
        _style_body_cell(c1, zebra=zebra)
        c2 = ws.cell(row=row, column=2, value=tname)
        _style_body_cell(c2, zebra=zebra)

        for j, did in enumerate(doc_ids, 3):
            val = row_data.get(did, 0)  # type: ignore[call-overload]
            c = ws.cell(row=row, column=j, value=val)
            c.font = BODY_FONT
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = BORDER
            c.fill = PRESENT_FILL if val == 1 else ABSENT_FILL

    # Summary row: "Темы на документ"
    summary_row = len(sorted_themes) + 2 + 1
    ws.cell(row=summary_row, column=1, value="Темы на документ").font = Font(
        name="Arial", size=10, bold=True, color="1F2937"
    )
    data_start = 2
    data_end = len(sorted_themes) + 1
    for j, did in enumerate(doc_ids, 3):
        col_letter = get_column_letter(j)
        c = ws.cell(
            row=summary_row,
            column=j,
            value=f"=COUNTIF({col_letter}{data_start}:{col_letter}{data_end},1)"
        )
        c.font = Font(name="Arial", size=10, bold=True, color="065A82")
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")

    _set_col_widths(ws, [10, 35] + [6] * 27)
    _freeze(ws, "C2")


# ---------------------------------------------------------------------------
# Sheet 06: Глубина покрытия по рангам
# ---------------------------------------------------------------------------

def build_sheet_06_coverage(wb: Workbook, data) -> None:
    ws = wb.create_sheet("06 Глубина покрытия по рангам")

    headers = ["ID темы", "Название", "Ранг 1", "Ранг 2", "Ранг 3", "Итого"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    # coverage_heatmap has: theme_id, theme_name, rank_1, rank_2, rank_3, rank_4
    def _theme_sort_key(row_data: dict) -> int:
        tid = row_data.get("theme_id", "")
        try:
            return int(tid[1:]) if tid.startswith("T") else 999
        except ValueError:
            return 999

    sorted_rows = sorted(data.coverage_heatmap, key=_theme_sort_key)

    for i, row_data in enumerate(sorted_rows):
        row = i + 2
        zebra = (i % 2 == 0)
        tid = row_data.get("theme_id", "")
        tname = row_data.get("theme_name", "")
        try:
            r1 = int(row_data.get("rank_1", 0) or 0)
            r2 = int(row_data.get("rank_2", 0) or 0)
            r3 = int(row_data.get("rank_3", 0) or 0)
        except ValueError:
            r1 = r2 = r3 = 0
        total = r1 + r2 + r3
        vals = [tid, tname, r1, r2, r3, total]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            _style_body_cell(c, zebra=zebra)

    # Conditional color scale on rank columns C:F
    n_rows = len(sorted_rows)
    if n_rows > 0:
        from openpyxl.formatting.rule import ColorScaleRule
        for col_letter in ["C", "D", "E", "F"]:
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{n_rows + 1}",
                ColorScaleRule(
                    start_type="min", start_color="F8FAFC",
                    end_type="max", end_color="1C7293"
                )
            )

    _set_col_widths(ws, [10, 40, 10, 10, 10, 10])
    _freeze(ws, "A2")


# ---------------------------------------------------------------------------
# Sheet 07: Граф зависимостей
# ---------------------------------------------------------------------------

def build_sheet_07_dependency_graph(wb: Workbook, data) -> None:
    ws = wb.create_sheet("07 Граф зависимостей")

    headers = ["От ID", "От — название", "К ID", "К — название", "Тип отношения", "Вес"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    for i, edge in enumerate(data.dependency_graph):
        row = i + 2
        zebra = (i % 2 == 0)
        rel_type = edge.get("relation_type", "")
        rel_fill = RELATION_FILLS.get(rel_type, ZEBRA_FILL if zebra else None)
        vals = [
            edge.get("from_doc_id", ""),
            edge.get("from_doc_short", ""),
            edge.get("to_doc_id", ""),
            edge.get("to_doc_short", ""),
            rel_type,
            edge.get("weight", ""),
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            if col == 5:
                _style_body_cell(c, fill=rel_fill)
            else:
                _style_body_cell(c, zebra=zebra)

    _set_col_widths(ws, [8, 35, 8, 35, 15, 8])
    _freeze(ws, "A2")


# ---------------------------------------------------------------------------
# Sheet 08: Матрица пар (351)
# ---------------------------------------------------------------------------

def build_sheet_08_pair_matrix(wb: Workbook, data) -> None:
    ws = wb.create_sheet("08 Матрица пар")

    headers = [
        "ID пары", "Левый ID", "Левый", "Правый ID", "Правый",
        "Темы", "Статус", "События", "Ранги"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    for i, pair in enumerate(data.pairs):
        row = i + 2
        zebra = (i % 2 == 0)
        status_eng = pair.get("v8_status", "").strip()
        status_display = _add_status_text(status_eng)
        left_id = pair.get("left", "")
        right_id = pair.get("right", "")
        # Get document titles
        left_doc = data.doc_by_id(left_id)
        right_doc = data.doc_by_id(right_id)
        left_title = (left_doc or {}).get("code", left_id)
        right_title = (right_doc or {}).get("code", right_id)

        vals = [
            pair.get("id", ""),
            left_id,
            left_title,
            right_id,
            right_title,
            pair.get("topics", ""),
            status_display,
            pair.get("events_count", ""),
            pair.get("rank_pair", ""),
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            if col == 7:
                _style_body_cell(c, status_eng=status_eng)
            else:
                _style_body_cell(c, zebra=zebra)

    _set_col_widths(ws, [12, 8, 28, 8, 28, 30, 26, 10, 10])
    _freeze(ws, "A2")


# ---------------------------------------------------------------------------
# Sheet 09: Все события (312)
# ---------------------------------------------------------------------------

def build_sheet_09_events(wb: Workbook, data) -> None:
    ws = wb.create_sheet("09 Все события")

    headers = [
        "ID события", "Тема", "Левый ID", "Левый", "Правый ID", "Правый",
        "Утверждение", "Доказательство", "Статус", "Уверенность"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    for i, ev in enumerate(data.events_all):
        row = i + 2
        zebra = (i % 2 == 0)
        status_eng = ev.get("status", "").strip()
        status_display = _add_status_text(status_eng)
        try:
            conf = float(ev.get("confidence", "") or 0)
        except (ValueError, TypeError):
            conf = ""  # type: ignore[assignment]

        vals = [
            ev.get("event_id", ""),
            ev.get("theme", ""),
            ev.get("left_id", ""),
            ev.get("left_doc", ""),
            ev.get("right_id", ""),
            ev.get("right_doc", ""),
            ev.get("claim_left", ""),
            ev.get("evidence_right", ""),
            status_display,
            conf,
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            if col == 9:
                _style_body_cell(c, status_eng=status_eng)
            elif col == 10 and isinstance(conf, float):
                _style_body_cell(c, zebra=zebra)
                c.number_format = "0.00"
            else:
                _style_body_cell(c, zebra=zebra)

    _set_col_widths(ws, [10, 25, 8, 22, 8, 22, 45, 45, 26, 10])
    _freeze(ws, "A2")


# ---------------------------------------------------------------------------
# Sheet 10: Тезисы НПА (87)
# ---------------------------------------------------------------------------

def build_sheet_10_theses(wb: Workbook, data) -> None:
    ws = wb.create_sheet("10 Тезисы НПА")

    headers = [
        "ID тезиса", "Тема", "Документ ID", "Документ",
        "Ранг", "Текст утверждения", "Источник", "Уверенность"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    for i, thesis in enumerate(data.theses):
        row = i + 2
        zebra = (i % 2 == 0)
        try:
            conf = float(thesis.get("confidence", "") or 0)
        except (ValueError, TypeError):
            conf = ""  # type: ignore[assignment]

        # Try to get rank from documents
        src_doc_name = thesis.get("source_doc", "")
        doc_id = thesis.get("source_doc", "")
        rank_val = ""
        # match by code in documents
        for doc in data.documents:
            if doc.get("code", "") == src_doc_name or doc.get("id", "") == doc_id:
                rank_val = doc.get("rank", "")
                doc_id = doc.get("id", doc_id)
                break

        vals = [
            thesis.get("thesis_id", ""),
            thesis.get("theme", ""),
            doc_id,
            thesis.get("source_doc", ""),
            rank_val,
            thesis.get("thesis", ""),
            thesis.get("coordinate", ""),
            conf,
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            _style_body_cell(c, zebra=zebra)
            if col == 8 and isinstance(conf, float):
                c.number_format = "0.00"

    _set_col_widths(ws, [10, 25, 8, 25, 6, 50, 30, 10])
    _freeze(ws, "A2")


# ---------------------------------------------------------------------------
# Sheet 11: Риски и противоречия (54)
# ---------------------------------------------------------------------------

def build_sheet_11_risks(wb: Workbook, data) -> None:
    ws = wb.create_sheet("11 Риски и противоречия")

    headers = [
        "ID риска", "Тема", "Статус", "Серьёзность",
        "Источник", "Описание проблемы", "Рекомендация", "Владелец"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    for i, risk in enumerate(data.risks):
        row = i + 2
        zebra = (i % 2 == 0)
        status_eng = risk.get("status", "").strip()
        status_display = _add_status_text(status_eng)
        severity = risk.get("risk", "").strip()

        vals = [
            risk.get("risk_id", ""),
            risk.get("theme", ""),
            status_display,
            severity,
            risk.get("source", ""),
            risk.get("issue", ""),
            risk.get("recommendation", ""),
            risk.get("owner", ""),
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            if col == 3:
                _style_body_cell(c, status_eng=status_eng)
            elif col == 4:
                sev_fill = SEVERITY_FILLS.get(severity.lower(), None)
                _style_body_cell(c, fill=sev_fill, zebra=zebra)
            else:
                _style_body_cell(c, zebra=zebra)

    _set_col_widths(ws, [10, 28, 26, 12, 30, 45, 45, 28])
    _freeze(ws, "A2")


# ---------------------------------------------------------------------------
# Sheet 12: Очередь ручной проверки (103)
# ---------------------------------------------------------------------------

def build_sheet_12_review_queue(wb: Workbook, data) -> None:
    ws = wb.create_sheet("12 Очередь ручной проверки")

    headers = [
        "ID задачи", "Приоритет", "ID события", "Тема",
        "Что проверить", "Зачем", "Источник",
        "Владелец", "Дедлайн", "Критерий закрытия", "Статус"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    for i, rq in enumerate(data.review_queue):
        row = i + 2
        zebra = (i % 2 == 0)
        priority = rq.get("priority", "").strip()
        p_fill = PRIORITY_FILLS.get(priority, None)

        vals = [
            rq.get("review_id", ""),
            priority,
            rq.get("event_id", ""),
            rq.get("theme", ""),
            rq.get("what_to_check", ""),
            rq.get("why", ""),
            rq.get("source", ""),
            rq.get("owner", ""),
            rq.get("deadline", ""),
            rq.get("closure_criteria", ""),
            rq.get("status", ""),
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            if col == 2:
                _style_body_cell(c, fill=p_fill, zebra=zebra)
            else:
                _style_body_cell(c, zebra=zebra)

    _set_col_widths(ws, [10, 10, 10, 28, 45, 40, 30, 25, 30, 40, 12])
    _freeze(ws, "A2")


# ---------------------------------------------------------------------------
# Sheet 13: Каталог действий (10)
# ---------------------------------------------------------------------------

def build_sheet_13_actions(wb: Workbook, data) -> None:
    ws = wb.create_sheet("13 Каталог действий")

    headers = [
        "ID действия", "Категория", "Серьёзность",
        "Где", "Что не так", "Что сделать", "Владелец"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    for i, action in enumerate(data.actions):
        row = i + 2
        zebra = (i % 2 == 0)
        severity = action.get("severity", "").strip()
        sev_fill = SEVERITY_FILLS.get(severity.lower(), None)

        vals = [
            action.get("id", ""),
            action.get("category", ""),
            severity,
            action.get("where", ""),
            action.get("what_is_wrong", ""),
            action.get("what_to_do", ""),
            action.get("owner", ""),
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            if col == 3:
                _style_body_cell(c, fill=sev_fill, zebra=zebra)
            else:
                _style_body_cell(c, zebra=zebra)

    _set_col_widths(ws, [12, 28, 12, 45, 45, 50, 35])
    _freeze(ws, "A2")


# ---------------------------------------------------------------------------
# Sheet 14: Provenance
# ---------------------------------------------------------------------------

def build_sheet_14_provenance(wb: Workbook, data) -> None:
    ws = wb.create_sheet("14 Provenance")

    headers = [
        "ID документа", "URL", "Статус загрузки",
        "Дата", "Fallback источник", "SHA-256", "Заметки"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    for i, prov in enumerate(data.provenance):
        row = i + 2
        zebra = (i % 2 == 0)
        status_raw = str(prov.get("status", "")).strip()
        # Color: copied/200 = green, error/4xx/5xx/timeout = red
        if status_raw in ("200", "ok", "copied", "downloaded"):
            status_fill = PatternFill("solid", fgColor="DCFCE7")
        elif any(status_raw.startswith(x) for x in ("4", "5", "timeout", "error", "fail")):
            status_fill = PatternFill("solid", fgColor="FEE2E2")
        else:
            status_fill = ZEBRA_FILL if zebra else None

        # Use manifest as doc id, title as url
        doc_id = prov.get("manifest", "")
        url = prov.get("url", "")
        sha = prov.get("sha256", "")
        local_path = prov.get("local_path", "")
        note = prov.get("note", "")
        title = prov.get("title", "")

        vals = [
            doc_id,
            url or title,
            status_raw,
            "",  # date not in CSV
            local_path,
            sha,
            note,
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            if col == 3:
                _style_body_cell(c, fill=status_fill)
            else:
                _style_body_cell(c, zebra=zebra)

    _set_col_widths(ws, [15, 45, 16, 12, 40, 40, 30])
    _freeze(ws, "A2")


# ---------------------------------------------------------------------------
# Sheet 15: QA-гейт
# ---------------------------------------------------------------------------

def build_sheet_15_qa(wb: Workbook, data) -> None:
    ws = wb.create_sheet("15 QA-гейт")

    # Banner header
    qa = data.qa
    passed = qa.get("passed", 0)
    total = qa.get("total", 0)
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = f"QA-гейт: {passed}/{total} PASS"
    c.font = Font(name="Arial", size=16, bold=True, color="166534")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor="DCFCE7")
    c.border = BORDER
    ws.row_dimensions[1].height = 36

    headers = ["Критерий", "Описание", "Статус", "Доказательство"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    _style_header_row(ws, 2, len(headers))

    checks = qa.get("checks", [])
    for i, check in enumerate(checks):
        row = i + 3
        zebra = (i % 2 == 0)
        status_str = check.get("status", "PASS")
        status_fill = QA_FILLS.get(status_str, None)

        vals = [
            check.get("name", ""),
            check.get("description", ""),
            status_str,
            check.get("evidence", ""),
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            if col == 3:
                _style_body_cell(c, fill=status_fill)
            else:
                _style_body_cell(c, zebra=zebra)

    _set_col_widths(ws, [12, 55, 10, 40])
    _freeze(ws, "A3")


# ---------------------------------------------------------------------------
# Sheet 16: Тренд итераций v7→v10
# ---------------------------------------------------------------------------

def build_sheet_16_trend(wb: Workbook, data) -> None:
    ws = wb.create_sheet("16 Тренд итераций")

    headers = [
        "Версия", "Дата", "Документов", "Пар", "Событий",
        "match", "partial", "contradiction", "outdated",
        "manual_review", "not_comparable",
        "Match-доля (%)", "Очередь проверки", "Источник"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    timeline = data.trend.get("timeline", [])
    for i, entry in enumerate(timeline):
        row = i + 2
        zebra = (i % 2 == 0)
        vals = [
            entry.get("version", ""),
            entry.get("date", "")[:10] if entry.get("date") else "",
            entry.get("docs", ""),
            entry.get("pairs", ""),
            entry.get("events", ""),
            entry.get("status_match", ""),
            entry.get("status_partial", ""),
            entry.get("status_contradiction", ""),
            entry.get("status_outdated", ""),
            entry.get("status_manual_review", ""),
            entry.get("status_not_comparable", ""),
            entry.get("match_share", ""),
            entry.get("review_queue", ""),
            entry.get("source", ""),
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            _style_body_cell(c, zebra=zebra)

    # Deltas section
    delta_row = len(timeline) + 3
    ws.cell(row=delta_row, column=1, value="Дельты версий").font = Font(
        name="Arial", size=13, bold=True, color="1F2937"
    )
    delta_row += 1

    delta_headers = ["От", "До", "Δ документов", "Δ match-доля", "Δ очередь", "Δ противоречий"]
    for col, h in enumerate(delta_headers, 1):
        ws.cell(row=delta_row, column=col, value=h)
    _style_header_row(ws, delta_row, len(delta_headers))
    delta_row += 1

    deltas = data.trend.get("deltas", [])
    for i, delta in enumerate(deltas):
        row = delta_row + i
        zebra = (i % 2 == 0)
        vals = [
            delta.get("from", ""),
            delta.get("to", ""),
            delta.get("docs_delta", ""),
            delta.get("match_share_delta", ""),
            delta.get("review_delta", ""),
            delta.get("contradiction_delta", ""),
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            _style_body_cell(c, zebra=zebra)

    _set_col_widths(ws, [8, 14, 12, 10, 10, 14, 12, 18, 12, 18, 18, 16, 18, 25])
    _freeze(ws, "A2")


# ---------------------------------------------------------------------------
# Sheet 17: Дельта v9→v10
# ---------------------------------------------------------------------------

def build_sheet_17_delta(wb: Workbook, data) -> None:
    ws = wb.create_sheet("17 Дельта v9→v10")

    # Banner
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "Дельта v9 → v10"
    c.font = Font(name="Arial", size=16, bold=True, color="1F2937")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor="EFF6FF")
    c.border = BORDER
    ws.row_dimensions[1].height = 36

    # Description
    ws.merge_cells("A2:E2")
    desc = ws["A2"]
    desc.value = (
        "v10 — рендеринг-релиз поверх v9. "
        "Корпус, события, пары без изменений. "
        "Новое: пояснительная записка (16 стр), редакционный diff, аналитические CSV."
    )
    desc.font = Font(name="Arial", size=10, italic=True, color="374151")
    desc.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 40

    # Control numbers table
    row = 4
    ws.cell(row=row, column=1, value="Контрольные числа").font = Font(
        name="Arial", size=13, bold=True, color="1F2937"
    )
    row += 1

    cn_headers = ["Метрика", "До (v9)", "После (v10)", "Δ"]
    for col, h in enumerate(cn_headers, 1):
        ws.cell(row=row, column=col, value=h)
    _style_header_row(ws, row, len(cn_headers))
    row += 1

    delta = data.delta
    cn = delta.get("control_numbers", {})
    cn_data = [
        ("Документов", cn.get("docs_old", ""), cn.get("docs_new", ""), cn.get("docs_new", 0) - cn.get("docs_old", 0) if isinstance(cn.get("docs_old"), int) else ""),
        ("Пар",        cn.get("pairs_old", ""), cn.get("pairs_new", ""), cn.get("pairs_new", 0) - cn.get("pairs_old", 0) if isinstance(cn.get("pairs_old"), int) else ""),
        ("Событий",    cn.get("events_old", ""), cn.get("events_new", ""), cn.get("events_new", 0) - cn.get("events_old", 0) if isinstance(cn.get("events_old"), int) else ""),
    ]
    for i, (metric, old, new, delta_v) in enumerate(cn_data):
        zebra = (i % 2 == 0)
        for col, val in enumerate([metric, old, new, delta_v], 1):
            c = ws.cell(row=row + i, column=col, value=val)
            _style_body_cell(c, zebra=zebra)
    row += len(cn_data) + 2

    # New artifacts table
    ws.cell(row=row, column=1, value="Новые артефакты v10").font = Font(
        name="Arial", size=13, bold=True, color="1F2937"
    )
    row += 1

    art_headers = ["#", "Артефакт"]
    for col, h in enumerate(art_headers, 1):
        ws.cell(row=row, column=col, value=h)
    _style_header_row(ws, row, len(art_headers))
    row += 1

    new_artifacts = delta.get("new_artifacts_added", [])
    for i, artifact in enumerate(new_artifacts):
        r = row + i
        zebra = (i % 2 == 0)
        c1 = ws.cell(row=r, column=1, value=i + 1)
        _style_body_cell(c1, zebra=zebra)
        c2 = ws.cell(row=r, column=2, value=artifact)
        _style_body_cell(c2, zebra=zebra)

    row += len(new_artifacts) + 2

    # Status distribution diff (if any)
    status_diff = delta.get("status_distribution_diff", {})
    if status_diff:
        ws.cell(row=row, column=1, value="Изменения в распределении статусов").font = Font(
            name="Arial", size=13, bold=True, color="1F2937"
        )
        row += 1
        sd_headers = ["Статус", "Δ"]
        for col, h in enumerate(sd_headers, 1):
            ws.cell(row=row, column=col, value=h)
        _style_header_row(ws, row, len(sd_headers))
        row += 1
        for i, (k, v) in enumerate(status_diff.items()):
            zebra = (i % 2 == 0)
            status_eng = k
            c1 = ws.cell(row=row + i, column=1, value=_add_status_text(status_eng))
            _style_body_cell(c1, status_eng=status_eng)
            c2 = ws.cell(row=row + i, column=2, value=v)
            _style_body_cell(c2, zebra=zebra)

    _set_col_widths(ws, [30, 50, 15, 10, 15])
    _freeze(ws, "A3")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def build_xlsx(out_path: Path | str | None = None, *, data=None) -> Path:
    """Build the consolidated 17-sheet XLSX. Returns the saved path.

    Mirrors the API of build_pptx/build_docx/build_html so the orchestrator
    can compose them uniformly.
    """
    if data is None:
        data = load_data()

    wb = Workbook()
    wb.remove(wb.active)  # remove default Sheet

    build_sheet_01_summary(wb, data)
    build_sheet_02_legend(wb, data)
    build_sheet_03_documents(wb, data)
    build_sheet_04_themes_catalog(wb, data)
    build_sheet_05_correlation(wb, data)
    build_sheet_06_coverage(wb, data)
    build_sheet_07_dependency_graph(wb, data)
    build_sheet_08_pair_matrix(wb, data)
    build_sheet_09_events(wb, data)
    build_sheet_10_theses(wb, data)
    build_sheet_11_risks(wb, data)
    build_sheet_12_review_queue(wb, data)
    build_sheet_13_actions(wb, data)
    build_sheet_14_provenance(wb, data)
    build_sheet_15_qa(wb, data)
    build_sheet_16_trend(wb, data)
    build_sheet_17_delta(wb, data)

    # Apply landscape + fit-to-width on every sheet so PDF prints clean
    from openpyxl.worksheet.page import PageMargins
    for name in wb.sheetnames:
        ws = wb[name]
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.print_options.horizontalCentered = True
        ws.page_margins = PageMargins(
            left=0.4, right=0.4, top=0.6, bottom=0.6, header=0.3, footer=0.3
        )
        if ws.max_row > 1:
            ws.print_title_rows = "1:1"

    out = Path(out_path) if out_path else OUT_PATH
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def main() -> None:
    print("Loading data...")
    data = load_data()
    print("Building workbook...")
    out = build_xlsx(data=data)
    from openpyxl import load_workbook
    wb = load_workbook(out)
    size_kb = out.stat().st_size // 1024
    print(f"Saved: {out} ({size_kb} KB), sheets={len(wb.sheetnames)}")
    print("Sheet names:")
    for name in wb.sheetnames:
        print(f"  {name}")


if __name__ == "__main__":
    main()
