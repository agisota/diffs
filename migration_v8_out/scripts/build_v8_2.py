#!/usr/bin/env python3
"""v8.2 cover-PDF + Excel-with-visuals + RACI sheet builder.

Produces (writes into /home/dev/diff/migration_v8_out/):
  docs/Forensic_v8_cover.pdf            — cover with heatmap + status pie
  docs/Несоответствия_и_действия.xlsx   — re-rendered with embedded heatmap
                                            in 00 README sheet + new 08 RACI sheet
  data/17_raci_matrix.csv               — RACI table for FA-01..FA-10
  logs/qa_v8_2.json
"""
from __future__ import annotations

import csv
import datetime as dt
import json
import sys
from pathlib import Path
from typing import Any

ROOT = Path("/home/dev/diff/migration_v8_out")
DATA = ROOT / "data"
DOCS = ROOT / "docs"
LOGS = ROOT / "logs"
VIS = DOCS / "visuals"

GENERATED_AT = dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%d %H:%M:%SZ")

# Re-use the system catalogue
sys.path.insert(0, "/home/dev/diff/docdiffops_mvp")
from docdiffops.forensic_actions import DEFAULT_ACTIONS, raci_for_action  # noqa: E402


# ---------------------------------------------------------------------------
# 1. RACI matrix CSV
# ---------------------------------------------------------------------------

def write_raci_csv() -> Path:
    rows = []
    for a in DEFAULT_ACTIONS:
        raci = raci_for_action(a.id)
        rows.append({
            "ID": a.id,
            "Категория": a.category,
            "Серьёзность": a.severity,
            "R (исполнитель)": raci["R"],
            "A (подписывает)": raci["A"],
            "C (консультирует)": raci["C"],
            "I (информируется)": raci["I"],
            "Что сделать": a.what_to_do,
            "Связанные документы": ", ".join(a.related_docs),
        })
    p = DATA / "17_raci_matrix.csv"
    with p.open("w", encoding="utf-8-sig", newline="") as fh:
        w = csv.DictWriter(
            fh,
            fieldnames=["ID", "Категория", "Серьёзность",
                        "R (исполнитель)", "A (подписывает)", "C (консультирует)",
                        "I (информируется)", "Что сделать", "Связанные документы"],
            delimiter=";",
        )
        w.writeheader()
        for row in rows:
            w.writerow(row)
    return p


# ---------------------------------------------------------------------------
# 2. Embed heatmap + add RACI sheet to supplementary Excel
# ---------------------------------------------------------------------------

def enhance_supplementary_xlsx() -> Path:
    """Re-open the supplementary workbook, embed heatmap on README, add RACI sheet."""
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    p = DOCS / "Несоответствия_и_действия.xlsx"
    wb = load_workbook(p)

    head_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    head_font = Font(bold=True, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")
    thin = Side(border_style="thin", color="D1D5DB")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Embed heatmap on README sheet
    if "00 README" in wb.sheetnames:
        ws = wb["00 README"]
        heatmap_path = VIS / "heatmap_doc_x_doc.png"
        if heatmap_path.exists():
            img = XLImage(str(heatmap_path))
            img.width, img.height = 720, 620
            anchor_row = ws.max_row + 3
            ws.cell(row=anchor_row, column=1, value="Doc × Doc heatmap (визуализация 26×26 матрицы):").font = Font(bold=True)
            ws.add_image(img, f"A{anchor_row + 1}")

    # Add RACI sheet
    raci_sheet_name = "08 RACI"
    if raci_sheet_name in wb.sheetnames:
        del wb[raci_sheet_name]
    ws = wb.create_sheet(raci_sheet_name)
    header = ["ID", "Категория", "Серьёзность",
              "R (исполнитель)", "A (подписывает)", "C (консультирует)",
              "I (информируется)", "Что сделать"]
    ws.append(header)
    for cell in ws[1]:
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = wrap
        cell.border = border
    severity_fill = {
        "high": PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid"),
        "medium": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "low": PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
    }
    for a in DEFAULT_ACTIONS:
        raci = raci_for_action(a.id)
        ws.append([
            a.id, a.category, a.severity,
            raci["R"], raci["A"], raci["C"], raci["I"],
            a.what_to_do,
        ])
        cell = ws.cell(row=ws.max_row, column=3)
        cell.fill = severity_fill.get(a.severity, severity_fill["medium"])
    widths = {"ID": 9, "Категория": 24, "Серьёзность": 14,
              "R (исполнитель)": 32, "A (подписывает)": 32,
              "C (консультирует)": 32, "I (информируется)": 32,
              "Что сделать": 60}
    for i, h in enumerate(header, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths[h]
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = wrap
            cell.border = border

    wb.save(p)
    return p


# ---------------------------------------------------------------------------
# 3. Cover-page PDF with embedded visuals
# ---------------------------------------------------------------------------

def render_cover_pdf() -> Path:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import (PageBreak, Paragraph, SimpleDocTemplate,
                                     Spacer, Table, TableStyle)
    from reportlab.platypus import Image as RLImage

    candidates = [
        ("/usr/share/fonts/noto/NotoSans-Regular.ttf",
         "/usr/share/fonts/noto/NotoSans-Bold.ttf"),
        ("/usr/share/fonts/liberation/LiberationSans-Regular.ttf",
         "/usr/share/fonts/liberation/LiberationSans-Bold.ttf"),
    ]
    base, bold = "Helvetica", "Helvetica-Bold"
    for r_path, b_path in candidates:
        if Path(r_path).exists() and Path(b_path).exists():
            try:
                pdfmetrics.registerFont(TTFont("V8Sans", r_path))
                pdfmetrics.registerFont(TTFont("V8Sans-Bold", b_path))
                base, bold = "V8Sans", "V8Sans-Bold"
                break
            except Exception:
                continue

    out = DOCS / "Forensic_v8_cover.pdf"
    doc = SimpleDocTemplate(str(out), pagesize=A4,
                            leftMargin=1.2 * cm, rightMargin=1.2 * cm,
                            topMargin=1.2 * cm, bottomMargin=1.2 * cm)
    styles = getSampleStyleSheet()
    body = ParagraphStyle("body", parent=styles["BodyText"], fontName=base, fontSize=10, leading=13)
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontName=bold, fontSize=22, leading=26, spaceAfter=12)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontName=bold, fontSize=14, leading=18, spaceAfter=8)
    sub = ParagraphStyle("sub", parent=body, fontSize=10, textColor=colors.HexColor("#4B5563"))
    elems: list[Any] = []

    elems.append(Paragraph("Forensic v8.2 — DocDiffOps", h1))
    elems.append(Paragraph(
        "Evidence-grade интегральное перекрестное сравнение корпуса миграционных документов РФ. "
        "Не является юридическим заключением.", sub))
    elems.append(Spacer(1, 6))
    elems.append(Paragraph(f"Сгенерировано: {GENERATED_AT}", sub))
    elems.append(Spacer(1, 8))

    # Embedded composite cover summary
    cover = VIS / "cover_summary.png"
    if cover.exists():
        elems.append(RLImage(str(cover), width=18 * cm, height=14 * cm))
    elems.append(PageBreak())

    elems.append(Paragraph("Карта доменов (26×26)", h2))
    hm = VIS / "heatmap_doc_x_doc.png"
    if hm.exists():
        elems.append(RLImage(str(hm), width=18 * cm, height=15 * cm))
    elems.append(PageBreak())

    elems.append(Paragraph("Тематическое покрытие", h2))
    tb = VIS / "topic_bar.png"
    if tb.exists():
        elems.append(RLImage(str(tb), width=18 * cm, height=12 * cm))

    elems.append(Spacer(1, 8))
    elems.append(Paragraph("Распределение по рангам источников", h2))
    rb = VIS / "rank_pair_bar.png"
    if rb.exists():
        elems.append(RLImage(str(rb), width=14 * cm, height=9 * cm))
    elems.append(PageBreak())

    # RACI summary table
    elems.append(Paragraph("RACI — кто за что отвечает", h2))
    rows = [["ID", "Категория", "Серьёзность", "R (исполнитель)", "A (подписывает)"]]
    for a in DEFAULT_ACTIONS:
        raci = raci_for_action(a.id)
        rows.append([a.id, a.category, a.severity, raci["R"], raci["A"]])
    t = Table(rows, colWidths=[1.4 * cm, 4 * cm, 2 * cm, 5.5 * cm, 5.5 * cm], repeatRows=1)
    t.setStyle(TableStyle([
        ("FONT", (0, 0), (-1, 0), bold, 9),
        ("FONT", (0, 1), (-1, -1), base, 8),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D1D5DB")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elems.append(t)

    doc.build(elems)
    return out


# ---------------------------------------------------------------------------
# 4. CLI offline rebuild — `python -m docdiffops.forensic_cli rebuild <bundle.json> --out <dir>`
# ---------------------------------------------------------------------------

def write_offline_cli() -> Path:
    """Write a tiny CLI wrapper to docdiffops/forensic_cli.py that rebuilds
    a v8 bundle from a JSON file without needing pipeline.py."""
    cli_text = '''"""Offline rebuild CLI for forensic v8 bundles.

Usage:
    python -m docdiffops.forensic_cli rebuild <bundle.json> --out <dir>

Reads a saved bundle JSON, applies the actions catalogue, and re-renders
all five artifacts (xlsx, explanatory docx, redgreen docx, summary pdf,
plus the bundle itself for completeness) into ``--out``.
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from .forensic_actions import apply_actions_to_bundle
from .forensic_render import (
    render_v8_docx_explanatory,
    render_v8_docx_redgreen,
    render_v8_pdf_summary,
    render_v8_xlsx,
)


def cmd_rebuild(args: argparse.Namespace) -> int:
    bundle_path = Path(args.bundle)
    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)

    bundle = json.loads(bundle_path.read_text(encoding="utf-8"))
    if args.with_actions:
        bundle = apply_actions_to_bundle(bundle)

    (out_dir / "bundle.json").write_text(
        json.dumps(bundle, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    render_v8_xlsx(bundle, out_dir / "forensic_v8.xlsx")
    render_v8_docx_explanatory(bundle, out_dir / "forensic_v8_explanatory.docx")
    render_v8_docx_redgreen(bundle, out_dir / "forensic_v8_redgreen.docx")
    render_v8_pdf_summary(bundle, out_dir / "forensic_v8_summary.pdf")

    print(f"rebuilt 5 artifacts under {out_dir}")
    return 0


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(prog="forensic_cli")
    sub = parser.add_subparsers(dest="cmd", required=True)
    rb = sub.add_parser("rebuild", help="Rebuild bundle artifacts from saved JSON")
    rb.add_argument("bundle", help="Path to bundle.json")
    rb.add_argument("--out", required=True, help="Output directory")
    rb.add_argument("--with-actions", action="store_true",
                    help="Apply v8.1 actions catalogue and RACI to the bundle")
    rb.set_defaults(func=cmd_rebuild)

    args = parser.parse_args(argv)
    return args.func(args)


if __name__ == "__main__":
    sys.exit(main())
'''
    p = Path("/home/dev/diff/docdiffops_mvp/docdiffops/forensic_cli.py")
    p.write_text(cli_text, encoding="utf-8")
    return p


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    raci = write_raci_csv()
    print(f"  ✓ {raci.relative_to(ROOT)}: {raci.stat().st_size:,}b")

    xlsx = enhance_supplementary_xlsx()
    print(f"  ✓ {xlsx.relative_to(ROOT)}: {xlsx.stat().st_size:,}b (heatmap embedded + RACI sheet)")

    cover = render_cover_pdf()
    print(f"  ✓ {cover.relative_to(ROOT)}: {cover.stat().st_size:,}b")

    cli = write_offline_cli()
    print(f"  ✓ {cli.relative_to(Path('/home/dev/diff/docdiffops_mvp'))}")

    qa = {
        "generated_at": GENERATED_AT,
        "schema": "v8.2",
        "added_artifacts": {
            "raci_csv": str(raci.relative_to(ROOT)),
            "supplementary_xlsx_with_heatmap_and_raci": str(xlsx.relative_to(ROOT)),
            "cover_pdf": str(cover.relative_to(ROOT)),
            "offline_cli_module": "docdiffops.forensic_cli",
        },
        "visuals": [str(p.relative_to(ROOT)) for p in sorted(VIS.iterdir())
                    if p.suffix == ".png"],
    }
    (LOGS / "qa_v8_2.json").write_text(
        json.dumps(qa, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
